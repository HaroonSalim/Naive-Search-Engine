#Using pandas
import pandas as pd

# File path
file_path = 'Cell_Phones_and_Accessories_5.xlsx'  # Update file path as needed

# An empty list to store the parsed data
data = []

# Chunk size
chunk_size = 1000

# Read the Excel file in chunks
for chunk in pd.read_excel(file_path, chunksize=chunk_size):
    # Convert each chunk to a list of dictionaries
    chunk_dict = chunk.to_dict(orient='records')
    data.extend(chunk_dict)

# Display the first 5 records
for i in range(5):
    print(data[i])




#using openpyxl
from openpyxl import load_workbook

# File path
file_path = 'Cell_Phones_and_Accessories_5.xlsx'  # Update file path as needed

# An empty list to store the parsed data
data = []

# Chunk size
chunk_size = 1000

# Load the Excel workbook
workbook = load_workbook(filename=file_path, read_only=True)

# Select the active worksheet
worksheet = workbook.active

# Read the Excel file in chunks
for row in worksheet.iter_rows(min_row=2, values_only=True, max_row=chunk_size+1):
    # Convert each row to a dictionary
    row_dict = {col_name: cell_value for col_name, cell_value in zip(worksheet[1], row)}
    data.append(row_dict)

    if len(data) >= chunk_size:
        break

# Display the first 5 records
for i in range(5):
    print(data[i])
